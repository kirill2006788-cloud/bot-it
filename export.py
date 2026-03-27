from flask import Blueprint, jsonify, request, send_file
from flask_login import login_required, current_user
from models import db, User, GameScore, UserAchievement, ToolUsage, Notification
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import csv
import io
from datetime import datetime
import zipfile

export_bp = Blueprint('export', __name__)

@export_bp.route('/api/export/profile', methods=['GET'])
@login_required
def export_profile():
    """Экспорт профиля пользователя"""
    try:
        format_type = request.args.get('format', 'json')  # json, pdf, excel
        
        if format_type == 'json':
            return export_profile_json()
        elif format_type == 'pdf':
            return export_profile_pdf()
        elif format_type == 'excel':
            return export_profile_excel()
        else:
            return jsonify({'error': 'Неподдерживаемый формат'}), 400
            
    except Exception as e:
        return jsonify({'error': 'Ошибка при экспорте профиля'}), 500

def export_profile_json():
    """Экспорт профиля в JSON"""
    try:
        # Получаем данные пользователя
        user_data = current_user.to_dict()
        
        # Статистика игр
        game_stats = db.session.query(
            db.func.count(GameScore.id).label('games_played'),
            db.func.sum(GameScore.score).label('total_score'),
            db.func.avg(GameScore.score).label('avg_score')
        ).filter_by(user_id=current_user.id).first()
        
        # Достижения
        achievements = UserAchievement.query.filter_by(user_id=current_user.id).all()
        
        # Использование инструментов
        tool_usage = ToolUsage.query.filter_by(user_id=current_user.id).all()
        
        # Уведомления
        notifications = Notification.query.filter_by(user_id=current_user.id).limit(50).all()
        
        export_data = {
            'export_info': {
                'exported_at': datetime.utcnow().isoformat(),
                'user_id': current_user.id,
                'username': current_user.username
            },
            'profile': user_data,
            'game_statistics': {
                'games_played': game_stats.games_played or 0,
                'total_score': game_stats.total_score or 0,
                'avg_score': round(game_stats.avg_score or 0, 2)
            },
            'achievements': [achievement.to_dict() for achievement in achievements],
            'tool_usage': [tool.to_dict() for tool in tool_usage],
            'recent_notifications': [notification.to_dict() for notification in notifications]
        }
        
        # Создаем файл в памяти
        output = io.StringIO()
        json.dump(export_data, output, indent=2, ensure_ascii=False)
        output.seek(0)
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'profile_{current_user.username}_{datetime.now().strftime("%Y%m%d")}.json'
        )
        
    except Exception as e:
        return jsonify({'error': 'Ошибка при экспорте в JSON'}), 500

def export_profile_pdf():
    """Экспорт профиля в PDF"""
    try:
        # Создаем PDF в памяти
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Стили
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Центрирование
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#667eea')
        )
        
        # Содержимое
        story = []
        
        # Заголовок
        story.append(Paragraph("IT Helper Bot - Профиль пользователя", title_style))
        story.append(Spacer(1, 20))
        
        # Информация о пользователе
        story.append(Paragraph("Информация о пользователе", heading_style))
        user_info = [
            ['Поле', 'Значение'],
            ['Имя пользователя', current_user.username],
            ['Email', current_user.email],
            ['Имя', current_user.first_name or 'Не указано'],
            ['Фамилия', current_user.last_name or 'Не указано'],
            ['Язык', current_user.language],
            ['Тема', current_user.theme],
            ['Премиум', 'Да' if current_user.is_premium else 'Нет'],
            ['Дата регистрации', current_user.created_at.strftime('%d.%m.%Y %H:%M')],
            ['Последний вход', current_user.last_login.strftime('%d.%m.%Y %H:%M') if current_user.last_login else 'Никогда']
        ]
        
        user_table = Table(user_info, colWidths=[2*inch, 3*inch])
        user_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(user_table)
        story.append(Spacer(1, 20))
        
        # Статистика игр
        story.append(Paragraph("Статистика игр", heading_style))
        
        game_stats = db.session.query(
            db.func.count(GameScore.id).label('games_played'),
            db.func.sum(GameScore.score).label('total_score'),
            db.func.avg(GameScore.score).label('avg_score')
        ).filter_by(user_id=current_user.id).first()
        
        game_info = [
            ['Параметр', 'Значение'],
            ['Игр сыграно', str(game_stats.games_played or 0)],
            ['Общий счет', str(game_stats.total_score or 0)],
            ['Средний счет', str(round(game_stats.avg_score or 0, 2))]
        ]
        
        game_table = Table(game_info, colWidths=[2*inch, 3*inch])
        game_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(game_table)
        story.append(Spacer(1, 20))
        
        # Достижения
        story.append(Paragraph("Достижения", heading_style))
        
        achievements = UserAchievement.query.filter_by(user_id=current_user.id).all()
        if achievements:
            achievement_data = [['Достижение', 'Дата получения', 'Очки']]
            for achievement in achievements:
                achievement_data.append([
                    achievement.achievement.name,
                    achievement.earned_at.strftime('%d.%m.%Y'),
                    str(achievement.points_earned)
                ])
            
            achievement_table = Table(achievement_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            achievement_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffc107')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightyellow),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(achievement_table)
        else:
            story.append(Paragraph("Достижения не найдены", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Подпись
        story.append(Paragraph(f"Экспорт создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        
        # Строим PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'profile_{current_user.username}_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
        
    except Exception as e:
        return jsonify({'error': 'Ошибка при экспорте в PDF'}), 500

def export_profile_excel():
    """Экспорт профиля в Excel"""
    try:
        # Создаем Excel файл в памяти
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Лист с информацией о пользователе
        ws_profile = workbook.active
        ws_profile.title = "Профиль"
        
        ws_profile['A1'] = "IT Helper Bot - Профиль пользователя"
        ws_profile['A1'].font = Font(size=16, bold=True)
        ws_profile.merge_cells('A1:B1')
        
        # Информация о пользователе
        profile_data = [
            ['Поле', 'Значение'],
            ['Имя пользователя', current_user.username],
            ['Email', current_user.email],
            ['Имя', current_user.first_name or 'Не указано'],
            ['Фамилия', current_user.last_name or 'Не указано'],
            ['Язык', current_user.language],
            ['Тема', current_user.theme],
            ['Премиум', 'Да' if current_user.is_premium else 'Нет'],
            ['Дата регистрации', current_user.created_at.strftime('%d.%m.%Y %H:%M')],
            ['Последний вход', current_user.last_login.strftime('%d.%m.%Y %H:%M') if current_user.last_login else 'Никогда']
        ]
        
        for row_num, row_data in enumerate(profile_data, 3):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_profile.cell(row=row_num, column=col_num, value=value)
                if row_num == 3:  # Заголовок
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # Автоширина колонок
        for column in ws_profile.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_profile.column_dimensions[column_letter].width = adjusted_width
        
        # Лист со статистикой игр
        ws_games = workbook.create_sheet("Игры")
        
        ws_games['A1'] = "Статистика игр"
        ws_games['A1'].font = Font(size=16, bold=True)
        ws_games.merge_cells('A1:B1')
        
        game_stats = db.session.query(
            db.func.count(GameScore.id).label('games_played'),
            db.func.sum(GameScore.score).label('total_score'),
            db.func.avg(GameScore.score).label('avg_score')
        ).filter_by(user_id=current_user.id).first()
        
        game_data = [
            ['Параметр', 'Значение'],
            ['Игр сыграно', game_stats.games_played or 0],
            ['Общий счет', game_stats.total_score or 0],
            ['Средний счет', round(game_stats.avg_score or 0, 2)]
        ]
        
        for row_num, row_data in enumerate(game_data, 3):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_games.cell(row=row_num, column=col_num, value=value)
                if row_num == 3:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
                    cell.alignment = center_alignment
        
        # Лист с достижениями
        ws_achievements = workbook.create_sheet("Достижения")
        
        ws_achievements['A1'] = "Достижения"
        ws_achievements['A1'].font = Font(size=16, bold=True)
        ws_achievements.merge_cells('A1:C1')
        
        achievements = UserAchievement.query.filter_by(user_id=current_user.id).all()
        
        achievement_data = [['Достижение', 'Дата получения', 'Очки']]
        for achievement in achievements:
            achievement_data.append([
                achievement.achievement.name,
                achievement.earned_at.strftime('%d.%m.%Y'),
                achievement.points_earned
            ])
        
        for row_num, row_data in enumerate(achievement_data, 3):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_achievements.cell(row=row_num, column=col_num, value=value)
                if row_num == 3:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
                    cell.alignment = center_alignment
        
        # Сохраняем файл
        workbook.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'profile_{current_user.username}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': 'Ошибка при экспорте в Excel'}), 500

@export_bp.route('/api/export/games', methods=['GET'])
@login_required
def export_games():
    """Экспорт данных об играх"""
    try:
        format_type = request.args.get('format', 'csv')
        
        if format_type == 'csv':
            return export_games_csv()
        elif format_type == 'excel':
            return export_games_excel()
        else:
            return jsonify({'error': 'Неподдерживаемый формат'}), 400
            
    except Exception as e:
        return jsonify({'error': 'Ошибка при экспорте игр'}), 500

def export_games_csv():
    """Экспорт игр в CSV"""
    try:
        games = GameScore.query.filter_by(user_id=current_user.id).order_by(GameScore.completed_at.desc()).all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Заголовки
        writer.writerow(['Тип игры', 'Счет', 'Уровень', 'Время (сек)', 'Дата'])
        
        # Данные
        for game in games:
            writer.writerow([
                game.game_type,
                game.score,
                game.level,
                game.time_spent,
                game.completed_at.strftime('%d.%m.%Y %H:%M')
            ])
        
        output.seek(0)
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'games_{current_user.username}_{datetime.now().strftime("%Y%m%d")}.csv'
        )
        
    except Exception as e:
        return jsonify({'error': 'Ошибка при экспорте в CSV'}), 500

def export_games_excel():
    """Экспорт игр в Excel"""
    try:
        games = GameScore.query.filter_by(user_id=current_user.id).order_by(GameScore.completed_at.desc()).all()
        
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Игры"
        
        # Заголовки
        headers = ['Тип игры', 'Счет', 'Уровень', 'Время (сек)', 'Дата']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for row_num, game in enumerate(games, 2):
            ws.cell(row=row_num, column=1, value=game.game_type)
            ws.cell(row=row_num, column=2, value=game.score)
            ws.cell(row=row_num, column=3, value=game.level)
            ws.cell(row=row_num, column=4, value=game.time_spent)
            ws.cell(row=row_num, column=5, value=game.completed_at.strftime('%d.%m.%Y %H:%M'))
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'games_{current_user.username}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': 'Ошибка при экспорте в Excel'}), 500

@export_bp.route('/api/export/full', methods=['GET'])
@login_required
def export_full_data():
    """Полный экспорт всех данных пользователя"""
    try:
        # Создаем ZIP архив
        buffer = io.BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Экспортируем профиль в JSON
            profile_data = get_profile_data()
            zip_file.writestr('profile.json', json.dumps(profile_data, indent=2, ensure_ascii=False))
            
            # Экспортируем игры в CSV
            games_csv = get_games_csv()
            zip_file.writestr('games.csv', games_csv)
            
            # Экспортируем достижения в JSON
            achievements_data = get_achievements_data()
            zip_file.writestr('achievements.json', json.dumps(achievements_data, indent=2, ensure_ascii=False))
            
            # Экспортируем использование инструментов в CSV
            tools_csv = get_tools_csv()
            zip_file.writestr('tools.csv', tools_csv)
            
            # Создаем README файл
            readme_content = f"""
IT Helper Bot - Полный экспорт данных
=====================================

Пользователь: {current_user.username}
Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}

Содержимое архива:
- profile.json - Информация о профиле
- games.csv - Статистика игр
- achievements.json - Достижения
- tools.csv - Использование инструментов

Этот архив содержит все ваши данные из IT Helper Bot.
Сохраните его в безопасном месте.
            """
            zip_file.writestr('README.txt', readme_content)
        
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'it_helper_bot_full_export_{current_user.username}_{datetime.now().strftime("%Y%m%d")}.zip'
        )
        
    except Exception as e:
        return jsonify({'error': 'Ошибка при полном экспорте'}), 500

def get_profile_data():
    """Получение данных профиля"""
    return {
        'user': current_user.to_dict(),
        'exported_at': datetime.utcnow().isoformat()
    }

def get_games_csv():
    """Получение CSV данных игр"""
    games = GameScore.query.filter_by(user_id=current_user.id).order_by(GameScore.completed_at.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Тип игры', 'Счет', 'Уровень', 'Время (сек)', 'Дата'])
    
    for game in games:
        writer.writerow([
            game.game_type,
            game.score,
            game.level,
            game.time_spent,
            game.completed_at.strftime('%d.%m.%Y %H:%M')
        ])
    
    return output.getvalue()

def get_achievements_data():
    """Получение данных достижений"""
    achievements = UserAchievement.query.filter_by(user_id=current_user.id).all()
    return {
        'achievements': [achievement.to_dict() for achievement in achievements],
        'exported_at': datetime.utcnow().isoformat()
    }

def get_tools_csv():
    """Получение CSV данных инструментов"""
    tools = ToolUsage.query.filter_by(user_id=current_user.id).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Инструмент', 'Количество использований', 'Последнее использование'])
    
    for tool in tools:
        writer.writerow([
            tool.tool_name,
            tool.usage_count,
            tool.last_used.strftime('%d.%m.%Y %H:%M')
        ])
    
    return output.getvalue()
